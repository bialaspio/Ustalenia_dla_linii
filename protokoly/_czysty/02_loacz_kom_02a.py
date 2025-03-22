import openpyxl, os, glob 
from openpyxl.styles.borders import Border, Side



# Load workbook


main_dir = os.getcwd()

list_files = glob.glob (main_dir+'\\out\\bez_format\\*.xlsx')
#print list_files
for nazwa in list_files:
	print nazwa
	wb = openpyxl.load_workbook(nazwa)
	nazwa_file = os.path.basename(nazwa)
	nazwa_path = os.path.dirname(nazwa)
	nazwa_path_dest = nazwa_path.replace('bez_format','format') +'\\'
	nazwa2 = '_'+nazwa_file
	
	print nazwa_file
	print nazwa_path
	print '--------'
	print nazwa_path_dest
	print nazwa2
	#os.system("pause")
	# Initialize formatting styles

	border_line = openpyxl.styles.borders.Border(left=Side(style='thin'), 
						 right=Side(style='thin'), 
						 top=Side(style='thin'), 
						 bottom=Side(style='thin')
						 )

	#cell_syle = openpyxl.styles.alignment.vertical ( Alignment.VERTICAL_BOTTOM )
	# Save workbook
	licznik = 1
	ws = wb.active
	for max_row, row in enumerate(ws, 1):
		a=1
	L_WIERSZY = max_row
	for sheet in wb.worksheets:
		for row in sheet:
			if licznik > 6:
				print str(licznik+1)+":"+str(licznik+1)
				pier_kom = 'A'+str(licznik+1)
				druga_kom = 'A'+str(licznik+2)
				#print pier_kom+':'+druga_kom
				ws.cell(row=licznik+1, column=7).value=''
				ws.cell(row=licznik+2, column=7).value=''
				ws.merge_cells(pier_kom+':'+druga_kom)
				pier_kom = 'G'+str(licznik+1)
				druga_kom = 'G'+str(licznik+2)
				#print pier_kom+':'+druga_kom
				ws.merge_cells(pier_kom+':'+druga_kom)
				pier_kom = 'H'+str(licznik+1)
				druga_kom = 'H'+str(licznik+2)
				#print pier_kom+':'+druga_kom
				ws.merge_cells(pier_kom+':'+druga_kom)
				pier_kom = 'I'+str(licznik+1)
				druga_kom = 'I'+str(licznik+2)
				#print pier_kom+':'+druga_kom
				ws.cell(row=licznik+1, column=10).value=''
				ws.cell(row=licznik+2, column=10).value=''
				ws.merge_cells(pier_kom+':'+druga_kom)
				pier_kom = 'J'+str(licznik+1)
				druga_kom = 'J'+str(licznik+2)
				#print pier_kom+':'+druga_kom
				ws.merge_cells(pier_kom+':'+druga_kom)
				licznik = licznik+1
			licznik=licznik+1	
			if licznik > L_WIERSZY:
				break			

	ws.column_dimensions['A'].width = 6.29
	ws.column_dimensions['B'].width = 21.43
	ws.column_dimensions['C'].width = 19.86
	ws.column_dimensions['D'].width = 21.86
	ws.column_dimensions['E'].width = 21.43
	ws.column_dimensions['F'].width = 11.43
	ws.column_dimensions['G'].width = 28.14
	ws.column_dimensions['H'].width = 11.86
	ws.column_dimensions['I'].width = 33
	ws.column_dimensions['J'].width = 27.86
	ws.column_dimensions['K'].width = 15.14


	licznik = 1
	for sheet in wb.worksheets:
		for row in sheet:
			if licznik > 6:
				print row 
				for cell in row:
					print cell
					cell.border = border_line
					cell
			licznik = licznik+1		
			if licznik > L_WIERSZY:
				break				
				
		
	wb.save(nazwa_path_dest+nazwa2)
	print nazwa


#zmiana pozycji w xls tam gdzie dzien i stanowisko (usuniecie gdziny )		

	
main_dir_format = main_dir+'\\out\\format\\*xlsx'
print main_dir_format 
files_xls_names = glob.glob(main_dir_format)
for file_xls_name in files_xls_names:	
	excelFile = openpyxl.load_workbook(file_xls_name, read_only=False)
	sh = excelFile.get_sheet_by_name('Sheet1')
	cell_value_g3 = sh["g3"].value
	print '--------------'
	print cell_value_g3
	print cell_value_g3[13:15] 
	print '--------------'
	text_replace = cell_value_g3[0:12]+' '+cell_value_g3[13:15] +cell_value_g3[15]
	sh["g3"].value = text_replace
	excelFile.save(file_xls_name)	
os.system("pause")